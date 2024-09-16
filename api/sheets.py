import logging
from django.forms import ValidationError
import openpyxl, xlsxwriter
import openpyxl.utils
import openpyxl.utils.dataframe

from .models import *
from .utils import FileManager, DataDumper

import pandas as pd 
import numpy as np
import time

from django.db import transaction
from django.db.models.aggregates import Sum, Count


from core.settings import GOOGLE_SHEET_CLIENT as gsclient

class DatabaseToGoogleSheetUpdater:

    # Partially Updating Only The Chapter in the sheet
    def __init__(self, episode_model: EpisodeModel | None, chapter_model: ChapterModel | None, reel_model: ReelModel | None, previousStartSequence, previousEndSequence, totalPreviousSequences = None, isDelete = False, isAdd = False):
        self.episode_model = episode_model
        self.chapter_model = chapter_model
        self.reel_model = reel_model
        self.isDeleted = isDelete
        self.isAdd = isAdd
        
        self.totalPreviousSequences = totalPreviousSequences
        self.previousStartSequence = previousStartSequence
        self.previousEndSequence = previousEndSequence

        self.sheet_link = episode_model.sheet_link

        # inital for updation
        self.chapter_filtered_sheet_reel_col_data = []
        self.chapter_sheet_dict = {
            "1": [],
            "Start Time": [],
            "End Time": [],
            "Text": [],
            "Chapter": [],
        }

        # self.chapters_sheet_dataframe = pd.DataFrame()

        # shortening the conditions
        if "episodes/" in episode_model.sheet_link:
            self.isFile = True
            if reel_model is not None:
                if isDelete:
                    self.get_reel_deleted_to_dataframe()
                elif isAdd:
                    self.get_reel_added_to_dataframe()
                else:
                    self.get_reel_to_dataframe()
                self.update_reel_file_sheet(self.sheet_link)
            elif chapter_model is not None:
                self.get_chapter_to_dataframe()
                self.update_chapter_file_sheet(self.sheet_link)
            elif episode_model is not None:
                self.get_episode_to_dataframe()
                self.update_episode_file_sheet(self.sheet_link)
        else:
            self.isFile = False
            if reel_model is not None:
                if isDelete:
                    self.get_reel_deleted_to_dataframe()
                elif isAdd:
                    self.get_reel_added_to_dataframe()
                else:
                    self.get_reel_to_dataframe()
                self.update_reel_google_sheet(self.sheet_link)
            elif chapter_model is not None:
                self.get_chapter_to_dataframe()
                self.update_chapter_google_sheet(self.sheet_link)
            elif episode_model is not None:
                self.get_episode_to_dataframe()
                self.update_episode_google_sheet(self.sheet_link)

    def get_reel_deleted_to_dataframe(self):

        print("DATABASE TO DATAFRAME")
        with transaction.atomic():
            start_time = time.time()
            
            chapter_model = self.chapter_model
            chapter_sequences = chapter_model.sequences.all()
            
            
            previous_reel_sequences_count = chapter_sequences.filter(sequence_number__gte = self.previousStartSequence, sequence_number__lte = self.previousEndSequence).count()
            
            chapters_before = ChapterModel.objects.filter(episode=self.episode_model, chapter_number__lt = chapter_model.chapter_number)
            total_sequences_before = sum([ch.sequences.all().count() for ch in chapters_before])
            total_sequences_before =  total_sequences_before + 1
            
            end_time = time.time()
            print(f"Fetched All Sequences, Chapters and Reels Data and Queryset -> {end_time - start_time} seconds.")
            
            # list preparation for partial updation            
            start_time = time.time()
            self.chapter_filtered_sheet_reel_col_data = np.full(previous_reel_sequences_count, "").tolist()
            startIndex = 0
            endIndex = 0
            totalIndex = 0
            for sequence in chapter_sequences.iterator(1500):
                if sequence.sequence_number == self.previousStartSequence:
                    startIndex = totalIndex
                    print(f"Prev Seq Found Start: ", startIndex)
                if sequence.sequence_number == self.previousEndSequence:
                    endIndex = totalIndex
                    print(f"Prev Seq Found End: ", startIndex)
                totalIndex += 1
                
            print(f"Reel Area Inside Chapter Sequence Indexes: ", startIndex, " ", endIndex, " ", total_sequences_before)
            self.chapter_filtered_reel_col_editable_range_start = total_sequences_before + startIndex
            self.chapter_filtered_reel_col_editable_range_end = total_sequences_before + endIndex
            
                    
            print(f"Chapter Filtered Sheet Range: ", self.chapter_filtered_reel_col_editable_range_start , " ", self.chapter_filtered_reel_col_editable_range_end)
            
            end_time = time.time()
            
            print(f"Chapters Filtered Reel Col List is Created -> {end_time - start_time} seconds.")

    
    def get_reel_added_to_dataframe(self):

        print("DATABASE TO DATAFRAME")
        with transaction.atomic():
            start_time = time.time()
            
            chapter_model = self.chapter_model
            chapter_sequences = chapter_model.sequences.all()
            
            reel_model = self.reel_model
            
            previous_reel_sequences_count = reel_model.sequences.all().count()
            
            chapters_before = ChapterModel.objects.filter(episode=self.episode_model, chapter_number__lt = chapter_model.chapter_number)
            total_sequences_before = sum([ch.sequences.all().count() for ch in chapters_before])
            total_sequences_before =  total_sequences_before + 1
            
            end_time = time.time()
            print(f"Fetched All Sequences, Chapters and Reels Data and Queryset -> {end_time - start_time} seconds.")
            
            # list preparation for partial updation            
            start_time = time.time()
            self.chapter_filtered_sheet_reel_col_data = np.full(previous_reel_sequences_count, str(reel_model.reel_number)).tolist()
            startIndex = 0
            endIndex = 0
            totalIndex = 0
            for sequence in chapter_sequences.iterator(1500):
                if sequence.sequence_number == self.previousStartSequence:
                    startIndex = totalIndex
                    print(f"Prev Seq Found Start: ", startIndex)
                if sequence.sequence_number == self.previousEndSequence:
                    endIndex = totalIndex
                    print(f"Prev Seq Found End: ", startIndex)
                totalIndex += 1
                
            print(f"Reel Area Inside Chapter Sequence Indexes: ", startIndex, " ", endIndex, " ", total_sequences_before)
            self.chapter_filtered_reel_col_editable_range_start = total_sequences_before + startIndex
            self.chapter_filtered_reel_col_editable_range_end = total_sequences_before + endIndex
            
                    
            print(f"Chapter Filtered Sheet Range: ", self.chapter_filtered_reel_col_editable_range_start , " ", self.chapter_filtered_reel_col_editable_range_end)
            
            end_time = time.time()
            
            print(f"Chapters Filtered Reel Col List is Created -> {end_time - start_time} seconds.")

            

    def get_reel_to_dataframe(self):

        print("DATABASE TO DATAFRAME")
        with transaction.atomic():
            start_time = time.time()
            
            chapter_model = self.chapter_model
            chapter_sequences = chapter_model.sequences.all()
            
            reel_model = self.reel_model
            reel_number = reel_model.reel_number
            
            previous_reel_sequences_count = chapter_sequences.filter(sequence_number__gte = self.previousStartSequence, sequence_number__lte = self.previousEndSequence).count()
            
            # reels_before = ReelModel.objects.filter(episode=self.episode_model, chapter=chapter_model, reel_number__lt = reel_model.reel_number).prefetch_related('sequences')
            # reels_after = ReelModel.objects.filter(episode=self.episode_model, chapter=chapter_model, reel_number__gt = reel_model.reel_number).prefetch_related('sequences')
            
            
            chapters_before = ChapterModel.objects.filter(episode=self.episode_model, chapter_number__lt = chapter_model.chapter_number)
            total_sequences_before = sum([ch.sequences.all().count() for ch in chapters_before])
            total_sequences_before =  total_sequences_before + 1
            
            # all_chapters = reels_before | ChapterModel.objects.filter(pk=chapter_model.id) | reels_after
            
            # sequence_querysets = [rl.sequences for rl in reels_before] + [previous_reel_sequences] + [rl.sequences for rl in reels_after]
            
            
            end_time = time.time()
            print(f"Fetched All Sequences, Chapters and Reels Data and Queryset -> {end_time - start_time} seconds.")

            # Build a dictionary for chapter and reel sequence ids
            start_time = time.time()
            # reel_sequence_dict = {
            #     f"{rl.reel_number}": set(str(seq.id) for seq in rl.sequences.all())
            #     for rl in ReelModel.objects.filter(episode=self.episode_model, chapter=chapter_model).prefetch_related('sequences')
            # }
            this_reel_sequences_numbers = set(seq.sequence_number for seq in reel_model.sequences.all())
            
            
            end_time = time.time()
            print(f"Reels Id Dictionary is Ready -> {end_time - start_time} seconds.")

            """
            # calculating editable range for chapter data in Chapters Sheet
            self.previousStartSequence = ch_sequences.first().sequence_number
            self.previousEndSequence = ch_sequences.last().sequence_number
            self.chapter_edit_range = f"D{self.previousStartSequence+1}:E{self.previousEndSequence+1}"
            print(self.chapter_edit_range)
            
            
            sequences = SequenceModel.objects.filter(sequence_number__gte = self.previousStartSequence, sequence_number__lte = self.previousEndSequence)
            total_sequences = sequences.count()

            # calculating reel ranges for Chapters Filtered Sheet
            # getting previous chapters number of sequences
            previous_sequences_count = sum([ech.sequences.get_queryset().count() for ech in ChapterModel.objects.filter(episode=self.episode_model, chapter_number__lt = chapter_model.chapter_number)])
            self.start_reel_row = previous_sequences_count + 2
            self.end_stop_row = previous_sequences_count + total_sequences + 1
            self.reel_col_edit_range = f"F{self.start_reel_row}:F{self.end_stop_row}"
            print(self.reel_col_edit_range)
            """
            
            # list preparation for partial updation            
            start_time = time.time()
            self.chapters_sheet_partial = []
            
            # creating array with present updates and overwriting the previous sequences with ''
            # so when the data will be updated the deleted sequences cells will get automatically to blank
            previous_reel_sequences = np.empty(previous_reel_sequences_count, dtype=SequenceModel)
            startIndex = 0
            endIndex = 0
            prevIndex = 0
            totalIndex = 0
            for sequence in chapter_sequences.iterator(1500):
                if sequence.sequence_number >= self.previousStartSequence and sequence.sequence_number <= self.previousEndSequence:
                    previous_reel_sequences[prevIndex] = sequence
                    prevIndex += 1
                if sequence.sequence_number == self.previousStartSequence:
                    startIndex = totalIndex
                    print(f"Prev Seq Found Start: ", startIndex)
                if sequence.sequence_number == self.previousEndSequence:
                    endIndex = totalIndex
                    print(f"Prev Seq Found End: ", startIndex)
                totalIndex += 1
                
            print(f"Reel Area Inside Chapter Sequence Indexes: ", startIndex, " ", endIndex, " ", total_sequences_before)
            self.chapter_filtered_reel_col_editable_range_start = total_sequences_before + startIndex
            self.chapter_filtered_reel_col_editable_range_end = total_sequences_before + endIndex
            
            currReelStartSequence = 0
            currReelEndSequence = 0
            reel_number_in_string = str(reel_number)
            for sequence in previous_reel_sequences:
                        # sequence.refresh_from_db()
                        # sequence_id = str(sequence.id)
                        sequence_num = sequence.sequence_number
                        reel_num = ""
                        
                        # for r_num, r_data in reel_sequence_dict.items():
                        if sequence_num in this_reel_sequences_numbers:
                                print(f"Sequence {sequence_num} found in Reel {reel_number_in_string}")
                                if currReelStartSequence == 0:
                                    currReelStartSequence = sequence_num
                                currReelEndSequence = sequence_num
                                reel_num = reel_number_in_string
                                

                        if currReelStartSequence > 0 and sequence.sequence_number <= self.previousEndSequence:
                            self.chapters_sheet_partial.append(sequence.words)
                        
                        # preparing array where sequence belong to chapters for chapter filtered reel column
                        self.chapter_filtered_sheet_reel_col_data.append(reel_num)
                    
            print(f"Chapter Sheet Range: ", currReelStartSequence , " ", currReelEndSequence)
            print(f"Chapter Filtered Sheet Range: ", self.chapter_filtered_reel_col_editable_range_start , " ", self.chapter_filtered_reel_col_editable_range_end)
            self.chapter_sheet_text_col_editable_range_start = currReelStartSequence
            self.chapter_sheet_text_col_editable_range_end = currReelEndSequence
            
            end_time = time.time()
            
            print(f"Chapters Sheet Partial List and Reel Col List is Created -> {end_time - start_time} seconds.")

            # self.chapters_sheet_dataframe = pd.DataFrame(self.chapter_sheet_dict)
        # print(f"Chapter Filtered Reel Col Size in MB", self.chapter_filtered_sheet_reel_col_data.__sizeof__() / 1024)
        # print(f"Chapter Sheet Partial Size in MB", self.chapters_sheet_partial.__sizeof__() / 1024)

    def update_reel_google_sheet(self, sheet_link):
        try:
            # return
            start_time = time.time()
            gexcel = gsclient.open_by_url(sheet_link)
            end_time = time.time()
            print(f"Google Sheet is Opened -> {end_time - start_time} seconds.")
            
            
            start_time = time.time()
            batch_updates = {
                "requests": [
                    {
                        "updateCells": {
                            "range": {
                                "sheetId": gexcel.worksheet('Chapter Filtered')._properties['sheetId'],
                                "startRowIndex": self.chapter_filtered_reel_col_editable_range_start,
                                "endRowIndex": self.chapter_filtered_reel_col_editable_range_end+1,
                                "startColumnIndex": 5,  # F column
                                "endColumnIndex": 6
                            },
                            "rows": [
                                {
                                    "values": [
                                        {
                                            "userEnteredValue": { 
                                                "stringValue": row
                                            }
                                        },
                                    ]
                                }
                                for row in self.chapter_filtered_sheet_reel_col_data
                            ],
                            "fields": "userEnteredValue"
                        }
                    },
                ]
            }
            
            if not self.isDeleted and not self.isAdd:
                batch_updates["requests"] = batch_updates["requests"] + [
                    {
                        "updateCells": {
                            "range": {
                                "sheetId": gexcel.worksheet('Copy CSV Here')._properties['sheetId'],
                                "startRowIndex": self.chapter_sheet_text_col_editable_range_start,
                                "endRowIndex": self.chapter_sheet_text_col_editable_range_end+1,
                                "startColumnIndex": 3,  # D column
                                "endColumnIndex": 4
                            },
                            "rows":[
                                {
                                    "values": [
                                        {
                                            "userEnteredValue": { 
                                                "stringValue": row
                                            }
                                        },
                                    ]
                                }
                                for row in self.chapters_sheet_partial
                            ],
                            "fields": "userEnteredValue"
                        }
                    },
                    {
                        "updateCells": {
                            "range": {
                                "sheetId": gexcel.worksheet('Chapters')._properties['sheetId'],  # Get sheet ID dynamically
                                "startRowIndex": self.chapter_sheet_text_col_editable_range_start,
                                "endRowIndex": self.chapter_sheet_text_col_editable_range_end+1,
                                "startColumnIndex": 3,  # D column
                                "endColumnIndex": 4
                            },
                            "rows": [
                                {
                                    "values": [
                                        {
                                            "userEnteredValue": { 
                                                "stringValue": row
                                            }
                                        } 
                                    ]
                                }
                                for row in self.chapters_sheet_partial
                            ],
                            "fields": "userEnteredValue"
                        }
                    }
                ]
            
            # Execute the batch update
            gexcel.batch_update(batch_updates)
            end_time = time.time()
            print(f"Batch Updated -> {end_time - start_time} seconds.")
            
            """
            start_time = time.time()
            chapters_worksheet = gexcel.worksheet("Chapters")
            chapters_worksheet.update(self.chapter_edit_range, self.chapters_sheet_partial)
            end_time = time.time()
            print(f"Chapters Sheet Updated -> {end_time - start_time} seconds.")

            start_time = time.time()
            filtered_chapters_worksheet = gexcel.worksheet("Chapter Filtered")
            # Update "Reel" column in "Chapter Filtered" sheet
            # filtered_chapters_worksheet.update(self.reel_col_edit_range, self.chapter_filtered_sheet_reel_col_data)
            filtered_chapters_worksheet.update("F2:F", self.chapter_filtered_sheet_reel_col_data)
            end_time = time.time()
            print(f"Chapter Filtered Sheet Reel Col Updated -> {end_time - start_time} seconds.")
            
            start_time = time.time()
            copy_csv_here_worksheet = gexcel.worksheet("Copy CSV Here")
            # Update "Text" column in "Copy CSV Here" sheet
            copy_csv_here_worksheet.update(f"D{(self.previousStartSequence+1)}:D{(self.previousEndSequence+1)}", [ [arr[0]] for arr in self.chapters_sheet_partial ])
            end_time = time.time()
            print(f"Copy CSV Here Sheet Text Col Updated -> {end_time - start_time} seconds.")
            """
        except Exception as ex:
            print(ex)
            logging.error(ex)
            raise ValidationError("Unable to download and update google sheet. Check Google API credentials or Network Connection.")            


    def update_reel_file_sheet(self, file_path):
        
        try:
            start_time = time.time()
            with FileManager.open(file_path, "rb") as file:
                workbook = openpyxl.load_workbook(file)
                
                end_time = time.time()
                print(f"Excel Sheet is Opened -> {end_time - start_time} seconds.")
                
                start_time = time.time()
                
                chapter_filtered_sheet = workbook["Chapter Filtered"]
                # Update "Reel" column in "Chapter Filtered" sheet
                for rowIndex, reel_value in enumerate(self.chapter_filtered_sheet_reel_col_data, self.chapter_filtered_reel_col_editable_range_start+1):
                    chapter_filtered_sheet[f"F{rowIndex}"].value = reel_value
                
                if not self.isDeleted and not self.isAdd:    
                    chapters_sheet = workbook["Chapters"]
                    # chapters_sheet_df = openpyxl.utils.dataframe.dataframe_to_rows(self.chapters_sheet_dataframe, index=False, header=True)
                    # print(chapters_sheet.print_title_rows)
                    for r_idx, rowData in enumerate(self.chapters_sheet_partial, self.chapter_sheet_text_col_editable_range_start+1):
                        chapters_sheet[f"D{r_idx}"] = rowData # Text Column Value

                    copy_csv_here_sheet = workbook["Copy CSV Here"]
                    # Update "Text" column in "Copy CSV Here" sheet
                    for rowIndex, rowData in enumerate(self.chapters_sheet_partial, self.chapter_sheet_text_col_editable_range_start+1):
                        copy_csv_here_sheet[f"D{rowIndex}"].value = rowData

                end_time = time.time()
                print(f"In Memory Modifications Done -> {end_time - start_time} seconds.")
                
                start_time = time.time()
                # Save the updated workbook
                with FileManager.open(file_path, "wb") as file:
                    workbook.save(file)
                end_time = time.time()
                print(f"Saving Excel Sheet -> {end_time - start_time} seconds.")
        except Exception as ex:
            print(ex)
            raise ValidationError("Unable to load and update excel file.")     

    def get_chapter_to_dataframe(self):

        print("DATABASE TO DATAFRAME")
        with transaction.atomic():
            start_time = time.time()
            
            chapter_model = self.chapter_model
            current_sequences = chapter_model.sequences.all()
            total_current_sequences = current_sequences.count()
            print(f"Total Current Sequences: {total_current_sequences}")
            
            previous_ch_sequences = SequenceModel.objects.filter(episode=self.episode_model, sequence_number__gte = self.previousStartSequence, sequence_number__lte = self.previousEndSequence)
            
            print(f"Total Previous Sequences: {self.totalPreviousSequences}")
            
            self.total_chapter_deleted_sequences = self.totalPreviousSequences - total_current_sequences
            print(f"Total Deleted Sequences: {self.total_chapter_deleted_sequences}")
            
            
            # chapters_before = ChapterModel.objects.filter(episode=self.episode_model, chapter_number__lt = chapter_model.chapter_number).prefetch_related('sequences')
            # chapters_after = ChapterModel.objects.filter(episode=self.episode_model, chapter_number__gt = chapter_model.chapter_number).prefetch_related('sequences')
            
            # # all_chapters = ChapterModel.objects.filter(episode=self.episode_model).prefetch_related('sequences')
            # all_chapters = chapters_before | ChapterModel.objects.filter(pk=chapter_model.id) | chapters_after
            
            
            # all sequences count of chapters before this chapter
            
            chapters_before_sequences = ChapterModel.objects.filter(
                episode=self.episode_model, chapter_number__lt = chapter_model.chapter_number
            ).annotate(
                sequences_count=Count("sequences")
            ).aggregate(
                total_sequences_of_all_chapter=Sum("sequences_count")
            )
            
            total_sequences_before_chapter = chapters_before_sequences["total_sequences_of_all_chapter"] or 0
            print(f"Total Sequences Before Chapter: {total_sequences_before_chapter}")
            
            self.chapter_filtered_reel_col_editable_range_start = total_sequences_before_chapter + 1
            
            if self.isFile:
                chapters_after_sequences = ChapterModel.objects.filter(
                    episode=self.episode_model, chapter_number__gt = self.chapter_model.chapter_number
                ).annotate(
                    sequences_count=Count("sequences")
                ).aggregate(
                    total_sequences_of_all_chapter=Sum("sequences_count")
                )
                
                self.total_sequences_after_chapter = chapters_after_sequences["total_sequences_of_all_chapter"] or 0
                print(f"Total Sequences After Chapter: {self.total_sequences_after_chapter}")
                
            # sequence_querysets = [ch.sequences for ch in chapters_before] + [previous_ch_sequences] + [ch.sequences for ch in chapters_after]
            # min_start_time = all_chapters.first().start_time
            # max_end_time = all_chapters.last().end_time
            # sequences = SequenceModel.objects.filter(episode=self.episode_model, start_time__gte = min_start_time, end_time__lte = max_end_time).iterator(1000)
            # all_chapter_reels = ReelModel.objects.filter(episode=self.episode_model, chapter=chapter_model).prefetch_related('sequences')
            
            end_time = time.time()
            print(f"Fetched All Sequences, Chapters and Reels Data and Queryset -> {end_time - start_time} seconds.")

            # Build a dictionary for chapter and reel sequence ids
            start_time = time.time()
            # chapter_sequence_dict = {
            #     f"{ch.chapter_number}": {
            #         "sequences": set(str(seq.id) for seq in (ch.sequences.all() if ch.id != chapter_model.id else chapter_model.sequences.all())),
            #         "reels": {
            #             f"{rl.reel_number}": set(str(seq.id) for seq in rl.sequences.all())
            #             for rl in ReelModel.objects.filter(episode=self.episode_model, chapter=ch).prefetch_related('sequences')
            #             # for rl in [arl for arl in all_chapter_reels if arl.chapter .chapter_number== ch.chapter_number]
            #         }
            #     }
            #     for ch in all_chapters
            # }
            chapter_sequence_dict = {
                "sequences": set(str(seq.id) for seq in current_sequences),
                "reels": {
                        f"{rl.reel_number}": set(str(seq.id) for seq in rl.sequences.all())        
                    for rl in ReelModel.objects.filter(episode=self.episode_model, chapter=chapter_model).prefetch_related('sequences')
                }
            }

            end_time = time.time()
            print(f"Chapter and Reels Id Dictionary is Ready -> {end_time - start_time} seconds.")


            """
            # calculating editable range for chapter data in Chapters Sheet
            self.previousStartSequence = ch_sequences.first().sequence_number
            self.previousEndSequence = ch_sequences.last().sequence_number
            self.chapter_edit_range = f"D{self.previousStartSequence+1}:E{self.previousEndSequence+1}"
            print(self.chapter_edit_range)
            
            
            sequences = SequenceModel.objects.filter(sequence_number__gte = self.previousStartSequence, sequence_number__lte = self.previousEndSequence)
            total_sequences = sequences.count()

            # calculating reel ranges for Chapters Filtered Sheet
            # getting previous chapters number of sequences
            previous_sequences_count = sum([ech.sequences.get_queryset().count() for ech in ChapterModel.objects.filter(episode=self.episode_model, chapter_number__lt = chapter_model.chapter_number)])
            self.start_reel_row = previous_sequences_count + 2
            self.end_stop_row = previous_sequences_count + total_sequences + 1
            self.reel_col_edit_range = f"F{self.start_reel_row}:F{self.end_stop_row}"
            print(self.reel_col_edit_range)
            """
            
            
            # list preparation for partial updation            
            start_time = time.time()
            self.chapters_sheet_partial = []
            for sequence in previous_ch_sequences.iterator(2000):
            # for sequence_querset in sequence_querysets:
                # print(sequence_querset)
                # for sequence in sequence_querset.all():
                #         sequence.refresh_from_db()
                        sequence_id = str(sequence.id)
                        chapter_number = ""
                        reel_number = ""
                        # for c_num, c_data in chapter_sequence_dict.items():
                        #     if sequence_id in c_data["sequences"]:
                        #         print(f"Sequence {sequence_id} found in Chapter {c_num}")
                        #         chapter_number = c_num
                        #         for r_num, r_data in c_data["reels"].items():
                        #             if sequence_id in r_data:
                        #                 print(f"Sequence {sequence_id} found in Reel {r_num}")
                        #                 reel_number = r_num
                        #                 break
                        #         break
                        
                        if sequence_id in chapter_sequence_dict["sequences"]:                            
                            chapter_number = str(chapter_model.chapter_number)
                            for r_num, r_data in chapter_sequence_dict["reels"].items():
                                if sequence_id in r_data:
                                    print(f"Sequence {sequence_id} found in Reel {r_num}")
                                    reel_number = r_num
                                    break
                        
                        # preparing array where sequence belong to chapters for chapter filtered reel column
                        if chapter_number != "":
                            self.chapter_filtered_sheet_reel_col_data.append([reel_number])
                        
                        if sequence.sequence_number >= self.previousStartSequence and sequence.sequence_number <= self.previousEndSequence:
                            self.chapters_sheet_partial.append([sequence.words, chapter_number])
                        
                        
            self.chapter_filtered_sheet_reel_col_data = self.chapter_filtered_sheet_reel_col_data
            
            end_time = time.time()
            # raise ValidationError("sdasd")
            print(f"Chapters Sheet Partial List and Reel Col List is Created -> {end_time - start_time} seconds.")
            
            # self.chapters_sheet_dataframe = pd.DataFrame(self.chapter_sheet_dict)
        # print(f"Chapter Filtered Reel Col Size in MB", self.chapter_filtered_sheet_reel_col_data.__sizeof__() / 1024)
        # print(f"Chapter Sheet Partial Size in MB", self.chapters_sheet_partial.__sizeof__() / 1024)

    def update_chapter_google_sheet(self, sheet_link):
        try:
            start_time = time.time()
            gexcel = gsclient.open_by_url(sheet_link)
            end_time = time.time()
            print(f"Google Sheet is Opened -> {end_time - start_time} seconds.")
            
            
            start_time = time.time()
            batch_updates = {
                "requests": [
                    {
                        "updateCells": {
                            "range": {
                                "sheetId": gexcel.worksheet('Chapters')._properties['sheetId'],  # Get sheet ID dynamically
                                "startRowIndex": self.previousStartSequence,
                                "endRowIndex": self.previousEndSequence+1,
                                "startColumnIndex": 3,
                                "endColumnIndex": 5
                            },
                            "rows": [
                                {
                                    "values": [
                                        {
                                            "userEnteredValue": { 
                                                "stringValue": row[0] 
                                            }
                                        },
                                        {
                                            "userEnteredValue": { 
                                                "stringValue": row[1] 
                                            }
                                        }  
                                    ]
                                }
                                for row in self.chapters_sheet_partial
                            ],
                            "fields": "userEnteredValue"
                        }
                    },
                    {
                        "updateCells": {
                            "range": {
                                "sheetId": gexcel.worksheet('Copy CSV Here')._properties['sheetId'],
                                "startRowIndex": self.previousStartSequence,
                                "endRowIndex": self.previousEndSequence+1,
                                "startColumnIndex": 3,  # D column
                                "endColumnIndex": 4
                            },
                            "rows":[
                                {
                                    "values": [
                                        {
                                            "userEnteredValue": { 
                                                "stringValue": row[0] 
                                            }
                                        },
                                    ]
                                }
                                for row in self.chapters_sheet_partial
                            ],
                            "fields": "userEnteredValue"
                        }
                    }
                ]
            }
            
            if self.total_chapter_deleted_sequences > 0:
                batch_updates["requests"].append(
                    {
                        "deleteRange": {
                            "range": {
                                "sheetId": gexcel.worksheet('Chapter Filtered')._properties['sheetId'],
                                "startRowIndex": self.chapter_filtered_reel_col_editable_range_start,
                                "endRowIndex": self.chapter_filtered_reel_col_editable_range_start + self.total_chapter_deleted_sequences,
                                "startColumnIndex": 5,  # F column
                                "endColumnIndex": 6
                            },
                            "shiftDimension": "ROWS"
                        }
                    }
                )
            
            batch_updates["requests"].append(
                    {
                        "updateCells": {
                            "range": {
                                "sheetId": gexcel.worksheet('Chapter Filtered')._properties['sheetId'],
                                "startRowIndex": self.chapter_filtered_reel_col_editable_range_start,
                                "endRowIndex": self.chapter_filtered_reel_col_editable_range_start + len(self.chapter_filtered_sheet_reel_col_data),
                                "startColumnIndex": 5,  # F column
                                "endColumnIndex": 6
                            },
                            "rows": [
                                {
                                    "values": [
                                        {
                                            "userEnteredValue": { 
                                                "stringValue": row[0] 
                                            }
                                        },
                                    ]
                                }
                                for row in self.chapter_filtered_sheet_reel_col_data
                            ],
                            "fields": "userEnteredValue"
                        }
                    }
            )
            
            
            # DataDumper.dump_to_file("request.json",batch_updates)
            
            # Execute the batch update
            gexcel.batch_update(batch_updates)
            end_time = time.time()
            print(f"Batch Updated -> {end_time - start_time} seconds.")
            
            """
            start_time = time.time()
            chapters_worksheet = gexcel.worksheet("Chapters")
            chapters_worksheet.update(self.chapter_edit_range, self.chapters_sheet_partial)
            end_time = time.time()
            print(f"Chapters Sheet Updated -> {end_time - start_time} seconds.")

            start_time = time.time()
            filtered_chapters_worksheet = gexcel.worksheet("Chapter Filtered")
            # Update "Reel" column in "Chapter Filtered" sheet
            # filtered_chapters_worksheet.update(self.reel_col_edit_range, self.chapter_filtered_sheet_reel_col_data)
            filtered_chapters_worksheet.update("F2:F", self.chapter_filtered_sheet_reel_col_data)
            end_time = time.time()
            print(f"Chapter Filtered Sheet Reel Col Updated -> {end_time - start_time} seconds.")
            
            start_time = time.time()
            copy_csv_here_worksheet = gexcel.worksheet("Copy CSV Here")
            # Update "Text" column in "Copy CSV Here" sheet
            copy_csv_here_worksheet.update(f"D{(self.previousStartSequence+1)}:D{(self.previousEndSequence+1)}", [ [arr[0]] for arr in self.chapters_sheet_partial ])
            end_time = time.time()
            print(f"Copy CSV Here Sheet Text Col Updated -> {end_time - start_time} seconds.")
            """
        except Exception as ex:
            print(ex)
            logging.error(ex)
            raise ValidationError("Unable to download and update google sheet. Check Google API credentials or Network Connection.")            

    def update_chapter_file_sheet(self, file_path):
       
        try:
            with FileManager.open(file_path, "rb") as file:
                workbook = openpyxl.load_workbook(file)

                chapters_sheet = workbook["Chapters"]
                print(chapters_sheet)
                # chapters_sheet_df = openpyxl.utils.dataframe.dataframe_to_rows(self.chapters_sheet_dataframe, index=False, header=True)
                # print(chapters_sheet.print_title_rows)
                for r_idx, rowData in enumerate(self.chapters_sheet_partial, self.previousStartSequence+1):
                    chapters_sheet[f"D{r_idx}"] = rowData[0] # Text Column Value
                    chapters_sheet[f"E{r_idx}"] = rowData[1] # Chapter Column Value

                chapter_filtered_sheet = workbook["Chapter Filtered"]
                editable_index_start = self.chapter_filtered_reel_col_editable_range_start + 1
                total_current_sequences = len(self.chapter_filtered_sheet_reel_col_data)
                
                moving_start_index = editable_index_start + self.total_chapter_deleted_sequences
                moving_end_index = moving_start_index + total_current_sequences + self.total_sequences_after_chapter
                
                # first delete the deleted cells
                if self.total_chapter_deleted_sequences > 0:
                    chapter_filtered_sheet.move_range(f"F{moving_start_index}:F{moving_end_index}", -self.total_chapter_deleted_sequences)
                    
                # Update "Reel" column in "Chapter Filtered" sheet
                for row, reel_value in enumerate(self.chapter_filtered_sheet_reel_col_data, editable_index_start):
                    chapter_filtered_sheet[f"F{row}"].value = reel_value[0]

                copy_csv_here_sheet = workbook["Copy CSV Here"]
                # Update "Text" column in "Copy CSV Here" sheet
                for row, rowData in enumerate(self.chapters_sheet_partial, self.previousStartSequence+1):
                    copy_csv_here_sheet[f"D{row}"].value = rowData[0]

                # Save the updated workbook
                with FileManager.open(file_path, "wb") as file:
                    workbook.save(file)
        except Exception as ex:
            print(ex)
            raise ValidationError("Unable to load and update excel file.")     

    def get_episode_to_dataframe(self):
        print("DATABASE TO DATAFRAME")
        with transaction.atomic():
            sequences = SequenceModel.objects.filter(episode=self.episode_model).iterator(1000)
            all_chapters = ChapterModel.objects.filter(episode=self.episode_model)
            all_reels = ReelModel.objects.filter(episode=self.episode_model).select_related('chapter')

            print("Fetched All Sequences, Chapters and Reels")

            # Build a dictionary for chapter and reel sequence ids
            chapter_sequence_dict = {
                    f"{ch.chapter_number}":{
                        "sequences": [seq.id for seq in ch.sequences.get_queryset()],
                        "reels": {
                            f"{rl.reel_number}": [seq.id for seq in rl.sequences.get_queryset()]
                        for rl in [ch_reel for ch_reel in all_reels if ch_reel.chapter == ch]}
                    } 
                    for ch in all_chapters}

            print("Chapter and Reels Id Dictionary is Ready")

            for sequence in sequences:
                sequence.refresh_from_db()
                # index += 1
                self.chapter_sheet_dict["1"].append("")
                self.chapter_sheet_dict["Start Time"].append(sequence.start_time)
                self.chapter_sheet_dict["End Time"].append(sequence.end_time)
                self.chapter_sheet_dict["Text"].append(sequence.words)

                chapter_number = ""
                reel_number = ""
                for c_num, c_data in chapter_sequence_dict.items():
                    if sequence.id in c_data["sequences"]:
                        print(f"Sequence {sequence.id} found in Chapter {c_num}")
                        chapter_number = c_num
                        for r_num, r_data in c_data["reels"].items():
                            if sequence.id in r_data:
                                print(f"Sequence {sequence.id} found in Reel {r_num}")
                                reel_number = r_num
                                break
                        break

                self.chapter_sheet_dict["Chapter"].append(chapter_number)

                if chapter_number != "":
                    self.chapter_filtered_sheet_reel_col_data.append(reel_number)

            print("Chapters Sheet Dict and Reel Col List is Created")

            """
                raise KeyError(23213)
                Build a dictionary for chapter and reel ranges
                chapter_number_dict = {}
                for ch in all_chapters:
                    ch.refresh_from_db()
                    c_f = ch.sequences.first().sequence_number
                    c_l = ch.sequences.last().sequence_number
                    # print(ch.title)
                    # print("Content: ", ch.content) 
                    # print("Start Sequences: ", c_f, " End Sequence: ", c_l)   
                    chapter_number_dict[f"{ch.chapter_number}"] = {
                        "range": range(c_f, c_l+1),
                        "reels": {}
                    }

                    for rl in ReelModel.objects.filter(episode=self.episode, chapter=ch):
                        rl.refresh_from_db()
                        chapter_number_dict[f"{ch.chapter_number}"]["reels"][f"{rl.reel_number}"] = range(rl.sequences.first().sequence_number, rl.sequences.last().sequence_number + 1)

                print(chapter_number_dict)
                
                # index = 0
                for sequence in sequences:
                    sequence.refresh_from_db()
                    # index += 1
                    self.chapter_sheet_dict["1"].append("")
                    self.chapter_sheet_dict["Start Time"].append(sequence.start_time)
                    self.chapter_sheet_dict["End Time"].append(sequence.end_time)
                    self.chapter_sheet_dict["Text"].append(sequence.words)

                    chapter_number = ""
                    reel_number = ""
                    for c_num, c_data in chapter_number_dict.items():
                        # if index <= 50:
                        #     print("Seq Num: ",sequence.sequence_number, " Range: ", c_data["range"])
                        if sequence.sequence_number in c_data["range"]:
                            chapter_number = c_num
                            for r_num, r_data in c_data["reels"].items():
                                if sequence.sequence_number in r_data:
                                    reel_number = r_num
                                    break
                            break

                    self.chapter_sheet_dict["Chapter"].append(chapter_number)

                    if chapter_number != "":
                        self.chapter_filtered_sheet_reel_col_data.append(reel_number)
                """

            self.chapters_sheet_dataframe = pd.DataFrame(self.chapter_sheet_dict)
        print(self.chapter_filtered_sheet_reel_col_data)
        print(self.chapters_sheet_dataframe.head(50))

    def update_episode_google_sheet(self, sheet_link):
        try:
            gexcel = gsclient.open_by_url(sheet_link)

            chapters_worksheet = gexcel.worksheet("Chapters")
            chapters_worksheet.clear()
            chapters_worksheet.insert_row(list(self.chapter_sheet_dict.keys()))
            cf_list = self.chapters_sheet_dataframe.values.tolist()
            chapters_worksheet.append_rows(cf_list)

            filtered_chapters_worksheet = gexcel.worksheet("Chapter Filtered")
            # filtered_chapters_worksheet.resize(1)
            # Update "Reel" column in "Chapter Filtered" sheet
            filtered_chapters_worksheet.update("F2:F", [ [reel_number] for reel_number in self.chapter_filtered_sheet_reel_col_data ])

            copy_csv_here_worksheet = gexcel.worksheet("Copy CSV Here")
            # Update "Text" column in "Copy CSV Here" sheet
            copy_csv_here_worksheet.update("D2:D", [ [words] for words in self.chapter_sheet_dict["Text"] ])
        except Exception as ex:
            raise ValidationError("Unable to download and update google sheet. Check Google API credentials or Network Connection.")            

    def update_episode_file_sheet(self, file_path):

        try:
            with FileManager.open(file_path, "rb") as file:
                workbook = openpyxl.load_workbook(file)

                chapters_sheet = workbook["Chapters"]
                print(chapters_sheet)
                chapters_sheet_df = openpyxl.utils.dataframe.dataframe_to_rows(self.chapters_sheet_dataframe, index=False, header=True)
                # print(chapters_sheet.print_title_rows)
                for r_idx, row in enumerate(chapters_sheet_df, 1):
                    for c_idx, value in enumerate(row, 1):
                        chapters_sheet.cell(row=r_idx, column=c_idx, value=value)

                chapter_filtered_sheet = workbook["Chapter Filtered"]
                # Update "Reel" column in "Chapter Filtered" sheet
                for row, reel_value in enumerate(self.chapter_filtered_sheet_reel_col_data):
                    chapter_filtered_sheet[f"F{row+2}"].value = reel_value

                copy_csv_here_sheet = workbook["Chapter Filtered"]
                # Update "Text" column in "Copy CSV Here" sheet
                for row, words in enumerate(self.chapter_sheet_dict["Text"]):
                    copy_csv_here_sheet[f"D{row+2}"].value = words

                # Save the updated workbook
                with FileManager.open(file_path, "wb") as file:
                    workbook.save(file)
        except Exception as ex:
            print(ex)
            raise ValidationError("Unable to load and update excel file.")     


class GoogleSheetProcessor:
    def __init__(self, sheet, video_url: str, isFile: bool):
        self.video_url = video_url

        start_time = time.time()
        if not isFile:
            self.sheet_url = sheet
            # self.sheet_key = self.filter_url(sheet)
            self.excel_file = self.download_google_sheet()
            self.all_sheets = self.load_sheets_from_google_sheet()
        else:
            self.excel_file = sheet
            self.filename = sheet._name
            self.sheet_url = self.filename
            self.all_sheets = self.load_sheets_from_file()
        end_time = time.time()
        print(f"Initialization time: {end_time - start_time} seconds")

        start_time = time.time()
        self.episode_title = self.extract_episode_title()
        end_time = time.time()
        print(f"Extract episode title time: {end_time - start_time} seconds")

    # def filter_url(self, sheet_sharing_url: str) -> str:
    #     start_time = time.time()
    #     if "https://docs.google.com/spreadsheets/d/" in sheet_sharing_url:
    #         result = sheet_sharing_url.replace("https://docs.google.com/spreadsheets/d/", "").split("/")[0]
    #     else:
    #         raise ValidationError("Invalid Google Sheets sharing link")
    #     end_time = time.time()
    #     print(f"Filter URL time: {end_time - start_time} seconds")
    #     return result

    def download_google_sheet(self) -> gspread.spreadsheet.Spreadsheet:
        start_time = time.time()
        # self.download_url = f'https://docs.google.com/spreadsheet/ccc?key={self.sheet_key}&output=xlsx'
        # response = requests.get(self.download_url)
        # self.filename = self.extract_filename(response)
        try:
            worksheets = gsclient.open_by_url(self.sheet_url)
            self.filename = worksheets.title
        except Exception as ex:
            raise ValidationError("Unable to download Google Sheets file.")
        end_time = time.time()
        print(f"Download Google Sheet time: {end_time - start_time} seconds")
        return worksheets

    # def extract_filename(self, response: requests.Response) -> str:
    #     start_time = time.time()
    #     content_disposition = response.headers.get('content-disposition')
    #     if content_disposition:
    #         filenames = re.findall(r'filename\*?=([^;]+)', content_disposition)
    #         if filenames:
    #             filename = filenames[-1].strip()
    #             if filename.startswith("UTF-8''"):
    #                 filename = filename[len("UTF-8''"):]
    #                 filename = requests.utils.unquote(filename)
    #             result = filename.strip('"')
    #         else:
    #             raise ValidationError("Google Sheets File Provided has no name separated by '[EPISODE TITLE] - [TOOL NAME] - [STATUS]'")
    #     else:
    #         raise ValidationError("Google Sheets File Provided has no name separated by '[EPISODE TITLE] - [TOOL NAME] - [STATUS]'")
    #     end_time = time.time()
    #     print(f"Extract filename time: {end_time - start_time} seconds")
    #     return result

    def load_sheets_from_google_sheet(self) -> dict[str: pd.DataFrame]:
        start_time = time.time()
        result = {}
        sheet = self.excel_file.worksheet("Copy CSV Here").get_all_values()
        columns = sheet.pop(0)
        result["Copy CSV Here"] = pd.DataFrame(sheet, columns=columns)
        sheet = self.excel_file.worksheet("Chapter Filtered").get_all_values()
        columns = sheet.pop(0)
        result["Chapter Filtered"] = pd.DataFrame(sheet, columns=columns)
        end_time = time.time()
        print(f"Load sheets from Spreadsheet time: {end_time - start_time} seconds")
        return result

    def load_sheets_from_file(self) -> dict[str: pd.DataFrame]:
        start_time = time.time()
        excel_file = pd.ExcelFile(self.excel_file)
        print(pd.read_excel(excel_file, sheet_name="Chapter Filtered", header=0))
        result = {sheet_name: pd.read_excel(excel_file, sheet_name=sheet_name, header=0) for sheet_name in excel_file.sheet_names if sheet_name in ["Copy CSV Here", "Chapter Filtered"]}
        end_time = time.time()
        print(f"Load sheets from file time: {end_time - start_time} seconds")
        return result

    def extract_episode_title(self) -> str:
        start_time = time.time()
        result = self.filename.split("-")[0].strip()
        end_time = time.time()
        print(f"Extract episode title time: {end_time - start_time} seconds")
        return result

    def save_full_episode_series_sequence(self):
        start_time = time.time()
        sequences_sheet = self.all_sheets.get("Copy CSV Here")
        chapters_filtered_sheet = self.all_sheets.get("Chapter Filtered")

        if sequences_sheet is None or chapters_filtered_sheet is None:
            raise ValidationError("Sheets Contain Invalid Episode Data!")

        print(chapters_filtered_sheet)
        
        sequences_sheet['Start Time'] = sequences_sheet['Start Time'].astype(str)
        sequences_sheet['End Time'] = sequences_sheet['End Time'].astype(str)
        sequences_sheet['Text'] = sequences_sheet['Text'].astype(str)

        chapters_filtered_sheet['Start Time'] = chapters_filtered_sheet['Start Time'].astype(str)
        chapters_filtered_sheet['End Time'] = chapters_filtered_sheet['End Time'].astype(str)
        chapters_filtered_sheet['Text'] = chapters_filtered_sheet['Text'].astype(str)
        chapters_filtered_sheet['Chapter'] = chapters_filtered_sheet['Chapter'].replace('', np.nan).fillna(0).astype(int)
        chapters_filtered_sheet['Reel'] = chapters_filtered_sheet['Reel'].replace('', np.nan).fillna(0).astype(int)

        combined_df = pd.merge(
            sequences_sheet,
            chapters_filtered_sheet[['Start Time', 'End Time', 'Text', 'Chapter', 'Reel']],
            on=['Start Time', 'End Time', 'Text'],
            how='outer'
        ).fillna({'Chapter': 0, 'Reel': 0}).astype({'Chapter': int, 'Reel': int})

        episode_model = EpisodeModel.objects.create(
            id=uuid.uuid4(),
            title=self.episode_title,
            content="",
            start_time=None,
            end_time=None,
            sheet_link=self.sheet_url,
            video_link=self.video_url
        )

        sequence_models = []
        chapters = {}
        reels = {}
        sequence_number = 0
        for row in combined_df.itertuples(index=False):
            sequence_uid = uuid.uuid4()
            sequence_number += 1
            sequence_model = SequenceModel(
                id=sequence_uid,
                episode=episode_model,
                words=row.Text,
                sequence_number=sequence_number,
                start_time=row._1,
                end_time=row._2
            )
            sequence_models.append(sequence_model)

            if row.Chapter:
                chapters.setdefault(row.Chapter, []).append(sequence_model)
            if row.Reel:
                reels.setdefault(f"{row.Chapter}-{row.Reel}", []).append(sequence_model)

        SequenceModel.objects.bulk_create(sequence_models)

        chapter_models = []
        for chapter_number, sequence_models in chapters.items():
            chapter_uid = uuid.uuid4()
            chapter_model = ChapterModel(
                id=chapter_uid,
                episode=episode_model,
                title=f"Chapter {chapter_number}",
                chapter_number=chapter_number,
                content=" ".join(combined_df.loc[combined_df['Chapter'] == chapter_number, 'Text']),
                start_time=combined_df.loc[combined_df['Chapter'] == chapter_number, 'Start Time'].min(),
                end_time=combined_df.loc[combined_df['Chapter'] == chapter_number, 'End Time'].max(),
            )
            chapter_model.save()
            chapter_models.append(chapter_model)
            chapter_model.sequences.set(sequence_models)

        reel_models = []
        for reel_key, sequence_models in reels.items():
            chapter_number, reel_number = reel_key.split("-")
            chapter_number = int(chapter_number)
            reel_number = int(reel_number)
            reel_uid = uuid.uuid4()
            reel_model = ReelModel(
                id=reel_uid,
                episode=episode_model,
                chapter=next((c for c in chapter_models if c.chapter_number == chapter_number), None),
                title=f"Reel {reel_number}",
                reel_number=reel_number,
                content=" ".join(combined_df.loc[(combined_df['Reel'] == reel_number) & (combined_df['Chapter'] == chapter_number), 'Text']),
                start_time=combined_df.loc[(combined_df['Reel'] == reel_number) & (combined_df['Chapter'] == chapter_number), 'Start Time'].min(),
                end_time=combined_df.loc[(combined_df['Reel'] == reel_number) & (combined_df['Chapter'] == chapter_number), 'End Time'].max(),
            )
            reel_model.save()
            reel_models.append(reel_model)
            reel_model.sequences.set(sequence_models)

        episode_model.content = " ".join(combined_df['Text'])
        episode_model.start_time = combined_df['Start Time'].min()
        episode_model.end_time = combined_df['End Time'].max()
        
        # saving excel file to media storagee
        if type(self.excel_file) is not gspread.spreadsheet.Spreadsheet:
            print("SAVING FILE")
            file_path = f"episodes/{episode_model.id}.xlsx"
            episode_model.sheet_link = file_path
            FileManager.save(file_path, self.excel_file)
        
        episode_model.save()

        end_time = time.time()
        print(f"Save full episode series sequence time: {end_time - start_time} seconds")

    
        
        
        
    def __str__(self) -> str:
        return str({sheet: df.shape for sheet, df in self.all_sheets.items()})
